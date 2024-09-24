import openpyxl

# Paths to your documents
document_one_path = 'Copy of UPDATE_Deep (Cluster 1).xlsx'
document_two_path = document_one_path
output_path = 'results.xlsx'  # If you want to save the combined result to a new file

# Sheet names
doc_two_sheet_name = 'Surface areas overview'
doc_one_sheet_name = 'GACP Calculation'
output_sheet_name = 'OutputSheetName'  # Name for the output sheet

def read_input_parameters(sheet):
    parameters = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
        parameters.append(row)
    return parameters

def write_parameters_and_extract_results(doc_one_wb, doc_one_ws, input_params):
    results = []
    for params in input_params:
        # Write parameters into Document One
        for col_num, value in enumerate(params, start=1):
            doc_one_ws.cell(row=2, column=col_num, value=value)  # Assuming data starts from row 2

        # Optionally save to trigger recalculations (if needed)
        doc_one_wb.save(document_one_path)

        # Extract results - change column numbers based on your specific result locations
        result = tuple(doc_one_ws.cell(row=2, column=col_num).value for col_num in range(len(params) + 1, len(params) + 4))
        results.append(result)
    return results

def main():
    # Open both documents
    doc_two_wb = openpyxl.load_workbook(document_two_path)
    doc_two_ws = doc_two_wb[doc_two_sheet_name]

    doc_one_wb = openpyxl.load_workbook(document_one_path)
    doc_one_ws = doc_one_wb[doc_one_sheet_name]

    # Read input parameters from Document Two
    input_parameters = read_input_parameters(doc_two_ws)
    print(input_parameters)
    # Write parameters to Document One and extract results
    # results = write_parameters_and_extract_results(doc_one_wb, doc_one_ws, input_parameters)

    # # Optionally, save results to a new document
    # output_wb = openpyxl.Workbook()
    # output_ws = output_wb.active
    # output_ws.title = output_sheet_name

    # for row_index, result in enumerate(results, start=1):
    #     for col_index, value in enumerate(result, start=1):
    #         output_ws.cell(row=row_index, column=col_index, value=value)

    # output_wb.save(output_path)

if __name__ == "__main__":
    main()