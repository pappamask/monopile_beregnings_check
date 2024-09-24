import pandas as pd
import openpyxl
from win32com.client import DispatchEx


path = r"C:\Users\tsla\GitHub\smallthings\BOHR\UPDATE_Deep (Cluster 1) - Copy.xlsx"
data_sheet = "Surface areas overview"
calc_sheet = 'GACP Calculation'

def just_open(filename):
    xlApp = DispatchEx('Excel.Application')
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    # xlBook.RefreshAll()
    # xlApp.CalculateUntilAsyncQueriesDone()
    xlBook.Save()
    xlBook.Close()

just_open(path)

print('reading excel...')
df = pd.read_excel(path, data_sheet, index_col=0)

print('creating sum column')
df['sum'] = df['J-tubes (immersed, coated)'] + df['Boat landing (immersed, coated)'] + df['Jacket (immersed, coated)']

df_values = df[['Piles (embedded, bare steel)', 'Piles (immersed, bare steel)', 'sum']]

df_results = pd.DataFrame()
print('iterating through rows')
for index, row in df_values.iterrows():
    print(f'calculating {index}')
    doc_calc = openpyxl.load_workbook(path)
    sheet_calc = doc_calc[calc_sheet]
    sheet_calc.cell(row=21, column=8, value=row.iloc[1])
    sheet_calc.cell(row=22, column=8, value=row.iloc[2])
    sheet_calc.cell(row=25, column=8, value=row.iloc[0])

    doc_calc.save(path)
    doc_calc.close()
    # just_open(path)

    df_result = pd.read_excel(path, calc_sheet, index_col=0, usecols = "B, H", nrows = 6, skiprows=43, header=None)
    df_result.rename(columns={df_result.columns[0]:index}, inplace=True)

    df_results = pd.concat([df_results, df_result], axis=1)

print(df_results)
